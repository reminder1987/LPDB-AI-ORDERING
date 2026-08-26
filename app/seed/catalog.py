from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import select

from app.core.database import SessionLocal
from app.models.category_db import (
    IngredientCategoryDB,
    ProductCategoryDB,
)
from app.models.ingredient_db import IngredientDB
from app.models.product_db import ProductDB


SOURCE_DIR = (
    Path.home()
    / "Downloads"
    / "PROYECTO LPDB API ORDER AGENT"
)

PRODUCT_SOURCE = SOURCE_DIR / "LPDB_MASTER_SPEC_v3.xlsx"
INGREDIENT_SOURCE = SOURCE_DIR / "LPDB_Recipe_Engine_v1.xlsx"


def read_product_categories() -> list[str]:
    workbook = openpyxl.load_workbook(
        PRODUCT_SOURCE,
        data_only=True,
    )

    worksheet = workbook["Productos_Maestro_v2"]

    categories = {
        row[0]
        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[0]
    }

    return sorted(categories)


def read_ingredient_categories() -> list[str]:
    workbook = openpyxl.load_workbook(
        INGREDIENT_SOURCE,
        data_only=True,
    )

    worksheet = workbook["Receta categorizada"]

    categories = {
        row[3]
        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[3]
    }

    return sorted(categories)


def read_ingredients() -> dict[str, str]:
    workbook = openpyxl.load_workbook(
        INGREDIENT_SOURCE,
        data_only=True,
    )

    worksheet = workbook["Receta categorizada"]

    ingredients: dict[str, str] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        ingredient = row[2]
        category = row[3]

        if ingredient and category:
            ingredients[ingredient] = category

    return ingredients


def read_products() -> list[dict]:
    workbook = openpyxl.load_workbook(
        PRODUCT_SOURCE,
        data_only=True,
    )

    worksheet = workbook["Productos_Maestro_v2"]

    products = []

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        category = row[0]
        name = row[1]
        price = row[2]

        if not category or not name or price is None:
            continue

        price_text = (
            str(price)
            .strip()
            .replace("$", "")
            .replace(",", "")
        )

        products.append(
            {
                "name": name,
                "category": category,
                "price": Decimal(price_text),
            }
        )

    return products


def get_or_create_product_categories(
    db,
    names: list[str],
) -> dict[str, ProductCategoryDB]:
    categories = {}

    for name in names:
        category = db.scalar(
            select(ProductCategoryDB).where(
                ProductCategoryDB.name == name
            )
        )

        if category is None:
            category = ProductCategoryDB(
                name=name,
            )
            db.add(category)
            db.flush()

        categories[name] = category

    return categories


def get_or_create_ingredient_categories(
    db,
    names: list[str],
) -> dict[str, IngredientCategoryDB]:
    categories = {}

    for name in names:
        category = db.scalar(
            select(IngredientCategoryDB).where(
                IngredientCategoryDB.name == name
            )
        )

        if category is None:
            category = IngredientCategoryDB(
                name=name,
            )
            db.add(category)
            db.flush()

        categories[name] = category

    return categories


def seed_categories() -> None:
    product_categories = read_product_categories()
    ingredient_categories = read_ingredient_categories()

    db = SessionLocal()

    try:
        get_or_create_product_categories(
            db,
            product_categories,
        )

        get_or_create_ingredient_categories(
            db,
            ingredient_categories,
        )

        db.commit()

    except Exception:
        db.rollback()
        raise

    finally:
        db.close()


def seed_ingredients() -> None:
    ingredients = read_ingredients()

    db = SessionLocal()

    try:
        category_by_name = {
            category.name: category
            for category in db.execute(
                select(IngredientCategoryDB)
            ).scalars()
        }

        for ingredient_name, category_name in ingredients.items():
            category = category_by_name.get(category_name)

            if category is None:
                raise ValueError(
                    f"Categoría de ingrediente no encontrada: "
                    f"{category_name}"
                )

            existing = db.scalar(
                select(IngredientDB).where(
                    IngredientDB.name == ingredient_name
                )
            )

            if existing is None:
                db.add(
                    IngredientDB(
                        name=ingredient_name,
                        category_id=category.id,
                    )
                )

        db.commit()

    except Exception:
        db.rollback()
        raise

    finally:
        db.close()


def seed_products() -> None:
    products = read_products()

    db = SessionLocal()

    try:
        category_by_name = {
            category.name: category
            for category in db.execute(
                select(ProductCategoryDB)
            ).scalars()
        }

        for product in products:
            category = category_by_name.get(
                product["category"]
            )

            if category is None:
                raise ValueError(
                    f"Categoría de producto no encontrada: "
                    f"{product['category']}"
                )

            existing = db.scalar(
                select(ProductDB).where(
                    ProductDB.name == product["name"]
                )
            )

            if existing is None:
                db.add(
                    ProductDB(
                        name=product["name"],
                        category_id=category.id,
                        price=product["price"],
                    )
                )

        db.commit()

    except Exception:
        db.rollback()
        raise

    finally:
        db.close()


if __name__ == "__main__":
    seed_categories()
    seed_ingredients()
    seed_products()

    print(
        "Categorías, ingredientes y productos "
        "cargados correctamente."
    )